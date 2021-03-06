from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pymssql
import os
import smtplib
import datetime


# ###################################################
# SQL Server 연결
# ###################################################
def SQLConnection():

    DBHost = os.environ['DBHost']
    DBUser = os.environ['DBUser']
    DBPassWord = os.environ['DBPassWord']
    DBName = os.environ['DBName']

    conn = pymssql.connect(server = DBHost, user = DBUser, password = DBPassWord, database = DBName) 
    
    return conn

# ###################################################
# SP 결과 값
# ###################################################
def SQLResultData():

    SQLConn = SQLConnection()
          
    try:
        Cursor = SQLConn.cursor()
        Cursor.execute(SP 명칭)
        Row = Cursor.fetchone() 
       
        if Row is not None:

            while Row:           
               
                DMCurrencyList = [str(Row[0]), Row[1], Row[2], int(Row[3]), int(Row[4])]
                                
                yield DMCurrencyList
                Row = Cursor.fetchone()     
        else:
            yield 0

    except pymssql.Error as e:
        print(e)
    finally:            
        SQLConn.close()
                 
# ###################################################
# SP Detail 결과값
# ###################################################
def SQLResultDetailData(UseName, CurrencyName):
    
    SQLConn = SQLConnection()

    try:    
        Cursor = SQLConn.cursor()              
        Params = (UseName, CurrencyName)
        Cursor.execute(SP명칭 @USE_YN = %s, @CurrencyName = %s",(Params))
        Row = Cursor.fetchone() 
            
        if Row is not None:

            while Row:           
                DMCurrencyDetailList = [str(Row[0]), int(Row[1]), int(Row[2]), Row[3],
                                        int(Row[4]), Row[5], Row[6], int(Row[7]), 
                                        Row[8], Row[9], int(Row[10]), int(Row[11]),
                                        float(Row[12])              
                                       ]           
      
                yield DMCurrencyDetailList
                Row = Cursor.fetchone()     
        else:
            yield 0

    except pymssql.Error as e:
        print(e)
    finally:            
        SQLConn.close()

# ###################################################
# Email 전송
# ###################################################
def MailSend(NewFilePath,NewFileName):

    Msg = MIMEMultipart()

    Msg['Subject'] = 메일제목
    Msg.attach(MIMEText(메일문구
                        , 'plain'))
    
    Attachment = open(NewFilePath,'rb')
    Part = MIMEBase('application', 'octet-stream')
    Part.set_payload(Attachment.read())
    encoders.encode_base64(Part)
    Part.add_header('Content-Disposition', "Attachment", filename=NewFileName)
    Msg.attach(Part)


    Sender_email = 받는사람 메일
    PassWord = 메일비밀번호
    Rec_email = ["메일명칭"]

    Server = smtplib.SMTP('smtp.gmail.com',587)
    Server.starttls()
    Server.login(Sender_email, PassWord)
    print("Login Success")
    Server.sendmail(Sender_email, Rec_email, Msg.as_string())
    print("Email has been sent to ", Rec_email)

###################################################
# Excel 행 조작
###################################################
def ExcelRowOpration(TargetSheet, TargetSheetMinRow, TargetSheetMaxRow, Load_Wb, 
                     BasePath, FileName, FileFormat, DetailSheet,
                     DetailSheetMinRow, DetailSheetMaxRow):

    global SQLResultData    
    SQLResultData = SQLResultData()
    SQLCheckType = next(SQLResultData)


    if SQLCheckType != 0 :

        # Excel 행 삭제
        TargetSheet.delete_rows(TargetSheetMinRow, TargetSheetMaxRow)  

        # Detail Excel 행 삭제
        DetailSheet.delete_rows(DetailSheetMinRow, DetailSheetMaxRow)
       
        # 시트명칭 지정
        # CreateSheet = Load_Wb.create_sheet()
        # CreateSheet.title = 'EventType'

        # 현재날짜, 시간
        CurrentDateTime = datetime.datetime.now()
        CurrentYear = CurrentDateTime.year

        # RAW DATA 저장
        for Raw in SQLResultData:
                        
            BaseDateHour = str(Raw[0])
            CurrencyName = Raw[1]
            UseName  = Raw[2]

            TargetSheet.append(Raw)

            BaseDate = str(CurrentYear) + "-" + BaseDateHour[0:5].replace("/", "-") 
            BaseHour = int(BaseDateHour[6:8])
    
            for Event in SQLResultDetailData(UseName, CurrencyName):
                
                # 일자, 시간 동일 경우에만 EventType시트에 저장
                if BaseDate == Event[0] and BaseHour == Event[1]:                    
                    DetailSheet.append(Event)
                    # CreateSheet.append(Event)
                             
        #  Excel 기준 시트 숨기기    
        #  TargetSheet.sheet_state = "hidden"
                       
        NewFilePath = BasePath + FileName + "_" + BaseDateHour.replace("/", "").replace(" ", "_") + FileFormat
        NewFileName = FileName + "_" + BaseDateHour.replace("/", "").replace(" ", "_") + FileFormat
   
        # Excel File 새 저장
        Load_Wb.save(NewFilePath)

        # Mail 전송
        MailSend(NewFilePath,NewFileName)

        print('Excel File ReSave Success')
    else:
        print("Not Exists Data")
        return

###################################################
# 기능 호출
###################################################
if __name__ == '__main__':

    # Excel File 오픈
    BasePath = "F:/Alert/Excel/"
    FileName = "Currency"
    FileFormat = ".xlsx"

    Load_Wb = load_workbook(BasePath + FileName + FileFormat)

    TargetSheet = Load_Wb['Sheet1']
    DetailSheet = Load_Wb['EventType']

    TargetSheetMinRow = TargetSheet.min_row + 1
    TargetSheetMaxRow = TargetSheet.max_row

    DetailSheetMinRow = DetailSheet.min_row + 1
    DetailSheetMaxRow = DetailSheet.max_row 
    
    # Excel 행 관련 작업 호출
    ExcelRowOpration(TargetSheet, TargetSheetMinRow, TargetSheetMaxRow, Load_Wb, 
                    BasePath, FileName, FileFormat, DetailSheet,
                    DetailSheetMinRow, DetailSheetMaxRow)
    
    

    
